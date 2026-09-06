"""Single-workbook Excel (.xlsx) I/O for Customer + Master Data.

One template/round-trip workbook with four sheets:
  * "Customers"        -> full customer columns
  * "Purchasing Size"  -> Name, Description
  * "Segment"          -> Name, Description
  * "TOP"              -> Name, Description

Pure functions only (no DB access) so they are trivially testable and reusable.
Cell values are normalised to trimmed strings so downstream validation matches
the existing CSV code path exactly.
"""
from __future__ import annotations

import io
from typing import Dict, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

CUSTOMER_SHEET = "Customers"
CUSTOMER_HEADERS = [
    "Customer Code", "Customer Name", "Segment", "Purchasing Size", "Area",
    "Status", "Bad Debt", "Bad Debt Nominal", "Address", "Latitude",
    "Longitude", "Phone", "WhatsApp", "PIC Name", "Payment Terms", "Credit Limit",
]

# Sheet title -> entity_type used by the sync engine / master collections.
MASTER_SHEETS = {
    "Purchasing Size": "purchasing_size",
    "Segment": "segment",
    "TOP": "top",
    "Area": "area",
}
MASTER_HEADERS = ["Name", "Description"]

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="1F5297")


def _cell_str(value) -> str:
    """Normalise any openpyxl cell value to a trimmed string."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, bool):
        return "Yes" if value else "No"
    return str(value).strip()


def _style_header(ws, ncols: int) -> None:
    for col in range(1, ncols + 1):
        c = ws.cell(row=1, column=col)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
    ws.freeze_panes = "A2"


def _autosize(ws, headers: List[str]) -> None:
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(12, min(40, len(h) + 6))


def _new_workbook() -> Workbook:
    wb = Workbook()
    # First sheet becomes Customers.
    ws = wb.active
    ws.title = CUSTOMER_SHEET
    ws.append(CUSTOMER_HEADERS)
    _style_header(ws, len(CUSTOMER_HEADERS))
    _autosize(ws, CUSTOMER_HEADERS)
    for title in MASTER_SHEETS:
        mws = wb.create_sheet(title)
        mws.append(MASTER_HEADERS)
        _style_header(mws, len(MASTER_HEADERS))
        _autosize(mws, MASTER_HEADERS)
    return wb


def _to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_template() -> bytes:
    """A ready-to-fill template with header rows + one example customer row."""
    wb = _new_workbook()
    ws = wb[CUSTOMER_SHEET]
    ws.append([
        "CUST-001", "PT Contoh Sejahtera", "Distributor", "Large", "Jakarta",
        "Active", "No", 0, "Jl. Contoh No. 1, Jakarta", -6.2, 106.8,
        "021-1234567", "0812-1111-2222", "Budi", "30 Days", 100000000,
    ])
    # Example master rows so the user sees the expected shape.
    for title, samples in (
        ("Purchasing Size", ["Enterprise", "Large", "Medium", "Small", "Micro"]),
        ("Segment", ["Distributor", "Retail", "Wholesale"]),
        ("TOP", ["Cash", "14 Days", "30 Days"]),
        ("Area", ["Jakarta", "Bandung", "Surabaya"]),
    ):
        mws = wb[title]
        for s in samples:
            mws.append([s, ""])
    return _to_bytes(wb)


def build_export(customers: List[dict], masters: Dict[str, List[dict]]) -> bytes:
    """Serialise Local-DB rows into the round-trippable workbook."""
    wb = _new_workbook()
    ws = wb[CUSTOMER_SHEET]
    for d in customers:
        ws.append([
            d.get("customer_code", ""),
            d.get("customer_name", ""),
            d.get("segment", ""),
            d.get("purchasing_size", ""),
            d.get("area", ""),
            d.get("status", ""),
            "Yes" if d.get("bad_debt") else "No",
            d.get("bad_debt_nominal", 0),
            d.get("address", ""),
            "" if d.get("latitude") is None else d.get("latitude"),
            "" if d.get("longitude") is None else d.get("longitude"),
            d.get("phone", ""),
            d.get("whatsapp", ""),
            d.get("pic_name", ""),
            d.get("payment_terms", ""),
            d.get("credit_limit", 0),
        ])
    for title, entity in MASTER_SHEETS.items():
        mws = wb[title]
        for m in masters.get(entity, []):
            mws.append([m.get("name", ""), m.get("description", "")])
    return _to_bytes(wb)


def _read_sheet(ws) -> List[Dict[str, str]]:
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return []
    keys = [_cell_str(h).lower() for h in header]
    out: List[Dict[str, str]] = []
    for raw in rows_iter:
        values = [_cell_str(v) for v in raw]
        if not any(values):
            continue  # skip fully-empty rows
        row = {keys[i]: (values[i] if i < len(values) else "") for i in range(len(keys)) if keys[i]}
        out.append(row)
    return out


def parse(content: bytes) -> Dict[str, List[Dict[str, str]]]:
    """Parse the workbook -> {'customers': [...], 'purchasing_size': [...], ...}.

    Missing sheets are tolerated (return empty lists). Raises ValueError on a
    corrupt / non-xlsx file so the API can surface a real error message.
    """
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001
        raise ValueError(f"File Excel tidak valid: {type(e).__name__}")

    result: Dict[str, List[Dict[str, str]]] = {"customers": []}
    titles = {t.lower(): t for t in wb.sheetnames}

    cust_title = titles.get(CUSTOMER_SHEET.lower())
    if cust_title:
        result["customers"] = _read_sheet(wb[cust_title])

    for sheet_title, entity in MASTER_SHEETS.items():
        real = titles.get(sheet_title.lower())
        result[entity] = _read_sheet(wb[real]) if real else []
    return result
