"""Backend tests for admin xlsx import/export endpoints.

Covers the review request:
  - /api/admin/template.xlsx returns valid, non-empty xlsx
  - /api/admin/export.xlsx returns valid xlsx containing 48 customer rows
  - /api/admin/import/xlsx/preview handles: (a) re-uploaded export → all skipped,
    (b) template → 1 create, (c) non-xlsx → 400 clean error
  - /api/admin/import/xlsx/commit with a ZZTEST customer creates the row, then
    cleans up via DELETE + trash purge.
"""
from __future__ import annotations

import io
import os
import pathlib

import openpyxl
import pytest
import requests
from dotenv import load_dotenv

# Load public backend URL from frontend/.env (that is the source of truth).
load_dotenv(pathlib.Path(__file__).resolve().parents[2] / "frontend" / ".env")
BASE_URL = (os.environ.get("EXPO_PUBLIC_BACKEND_URL") or "").rstrip("/")
assert BASE_URL, "EXPO_PUBLIC_BACKEND_URL must be set (frontend/.env)"

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# ---------------------------------------------------------------------------
# Shared fixtures
# ---------------------------------------------------------------------------


@pytest.fixture(scope="module")
def admin_token() -> str:
    r = requests.post(
        f"{BASE_URL}/api/admin/login",
        json={"username": "admin", "password": "admin123"},
        timeout=20,
    )
    assert r.status_code == 200, f"admin login failed: {r.status_code} {r.text}"
    return r.json()["access_token"]


@pytest.fixture(scope="module")
def auth(admin_token: str) -> dict:
    return {"Authorization": f"Bearer {admin_token}"}


@pytest.fixture(scope="module")
def template_bytes(auth: dict) -> bytes:
    r = requests.get(f"{BASE_URL}/api/admin/template.xlsx", headers=auth, timeout=30)
    assert r.status_code == 200
    return r.content


@pytest.fixture(scope="module")
def export_bytes(auth: dict) -> bytes:
    r = requests.get(f"{BASE_URL}/api/admin/export.xlsx", headers=auth, timeout=60)
    assert r.status_code == 200
    return r.content


# ---------------------------------------------------------------------------
# Download tests (template + export)
# ---------------------------------------------------------------------------


class TestDownloads:
    def test_template_is_valid_xlsx(self, template_bytes: bytes):
        assert len(template_bytes) > 1000, "template unexpectedly small"
        wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
        # openpyxl parsing without exception is proof of validity.
        assert wb.sheetnames, "template has no sheets"

    def test_export_contains_48_customers(self, export_bytes: bytes):
        assert len(export_bytes) > 1000, "export unexpectedly small"
        wb = openpyxl.load_workbook(io.BytesIO(export_bytes))
        # The customer sheet name may be "Customer" or similar — pick the one
        # containing an ID/customer_code column and count non-empty data rows.
        cust_sheet = None
        for name in wb.sheetnames:
            ws = wb[name]
            header = [str(c.value or "").strip().lower() for c in ws[1]]
            if any(h in ("customer code", "customer_code", "code") for h in header):
                cust_sheet = ws
                break
        assert cust_sheet is not None, f"no customer sheet found in {wb.sheetnames}"
        # Count rows where the first cell (customer_code) is non-empty.
        data_rows = [
            row for row in cust_sheet.iter_rows(min_row=2, values_only=True)
            if row and row[0] not in (None, "")
        ]
        assert len(data_rows) == 48, f"expected 48 customer rows, got {len(data_rows)}"


# ---------------------------------------------------------------------------
# Preview tests
# ---------------------------------------------------------------------------


class TestPreview:
    def test_preview_export_all_skip(self, auth: dict, export_bytes: bytes):
        files = {"file": ("export.xlsx", export_bytes, XLSX_MIME)}
        r = requests.post(
            f"{BASE_URL}/api/admin/import/xlsx/preview",
            headers=auth,
            files=files,
            timeout=60,
        )
        assert r.status_code == 200, r.text
        body = r.json()
        counts = body["customers"]["counts"]
        assert counts["total"] == 48, counts
        assert counts["skip"] == 48, counts
        assert counts["create"] == 0
        assert counts["update"] == 0
        assert counts["error"] == 0

    def test_preview_template_one_create(self, auth: dict, template_bytes: bytes):
        files = {"file": ("template.xlsx", template_bytes, XLSX_MIME)}
        r = requests.post(
            f"{BASE_URL}/api/admin/import/xlsx/preview",
            headers=auth,
            files=files,
            timeout=60,
        )
        assert r.status_code == 200, r.text
        counts = r.json()["customers"]["counts"]
        assert counts["total"] == 1, counts
        assert counts["create"] == 1, counts

    def test_preview_non_xlsx_returns_400(self, auth: dict):
        files = {"file": ("notes.txt", b"this is not an xlsx", "text/plain")}
        r = requests.post(
            f"{BASE_URL}/api/admin/import/xlsx/preview",
            headers=auth,
            files=files,
            timeout=30,
        )
        assert r.status_code == 400, f"expected 400, got {r.status_code} {r.text}"
        body = r.json()
        assert "detail" in body and isinstance(body["detail"], str)


# ---------------------------------------------------------------------------
# Commit test with a ZZTEST customer + cleanup
# ---------------------------------------------------------------------------


def _build_single_customer_xlsx(template_bytes: bytes, code: str) -> bytes:
    """Take the template and put ONE customer with the given code into it."""
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    # Find customer sheet by header
    cust_sheet = None
    header: list[str] = []
    for name in wb.sheetnames:
        ws = wb[name]
        row = [str(c.value or "").strip() for c in ws[1]]
        if any(h.lower() in ("customer code", "customer_code") for h in row):
            cust_sheet = ws
            header = row
            break
    assert cust_sheet is not None, "template has no customer sheet"

    def col(name: str) -> int:
        target = name.lower().replace("_", " ")
        for i, h in enumerate(header, start=1):
            if h.lower().replace("_", " ") == target:
                return i
        raise KeyError(name)

    # The template ships with 1 sample row at row 2. Replace its code + name so
    # it becomes our ZZTEST customer, while keeping other fields consistent
    # (segment/purchasing_size/area/top values referenced by master sheets).
    cust_sheet.cell(row=2, column=col("customer_code"), value=code)
    cust_sheet.cell(row=2, column=col("customer_name"), value=f"TEST Import {code}")
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


class TestCommit:
    zztest_code = "ZZTEST001"
    created_id: str | None = None

    def test_commit_creates_customer(self, auth: dict, template_bytes: bytes):
        payload = _build_single_customer_xlsx(template_bytes, self.zztest_code)
        files = {"file": ("zztest.xlsx", payload, XLSX_MIME)}
        r = requests.post(
            f"{BASE_URL}/api/admin/import/xlsx/commit",
            headers=auth,
            files=files,
            timeout=60,
        )
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["customers"]["created"] == 1, body
        assert body["customers"]["failed"] == 0, body
        assert "queue_pending" in body

    def test_verify_customer_exists_and_cleanup(self, auth: dict):
        # Search for the imported customer
        r = requests.get(
            f"{BASE_URL}/api/customers?search={self.zztest_code}",
            headers=auth,
            timeout=30,
        )
        assert r.status_code == 200, r.text
        rows = r.json()
        match = next((c for c in rows if c["customer_code"] == self.zztest_code), None)
        assert match is not None, f"customer {self.zztest_code} not found; got {rows}"
        cid = match["id"]

        # DELETE (soft-delete moves to trash)
        d = requests.delete(
            f"{BASE_URL}/api/customers/{cid}", headers=auth, timeout=30
        )
        assert d.status_code in (200, 204), d.text

        # Purge from trash so it does not pollute seed data
        p = requests.post(
            f"{BASE_URL}/api/admin/trash/customer/purge",
            headers=auth,
            json={"ids": [cid]},
            timeout=30,
        )
        assert p.status_code == 200, p.text
        assert p.json().get("success") is True

        # Confirm it is gone
        gone = requests.get(
            f"{BASE_URL}/api/customers?search={self.zztest_code}",
            headers=auth,
            timeout=30,
        )
        assert gone.status_code == 200
        remaining = [c for c in gone.json() if c["customer_code"] == self.zztest_code]
        assert remaining == [], f"cleanup failed, still present: {remaining}"
