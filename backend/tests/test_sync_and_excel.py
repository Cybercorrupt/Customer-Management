"""End-to-end backend regression tests for:
 * one-way sync engine (queue behavior + tombstones + cancel)
 * master data CRUD (purchasing_size / segment / top)
 * user CRUD + login safety
 * logo upload + file serving
 * Excel template / export / import round-trip (single workbook)

Local DB is the source of truth. Supabase pushes may fail because the user has
not run the schema migration yet — that is out of scope for these tests.
"""
from __future__ import annotations

import io
import os
import time
from uuid import uuid4

import pytest
import requests
from openpyxl import Workbook, load_workbook
from pymongo import MongoClient

BASE = os.environ["EXPO_PUBLIC_BACKEND_URL"].rstrip("/") if os.environ.get(
    "EXPO_PUBLIC_BACKEND_URL"
) else "https://crm-hub-564.preview.emergentagent.com"
API = f"{BASE}/api"
MONGO_URL = "mongodb://localhost:27017"
DB_NAME = "test_database"


# ------------------------------ fixtures -----------------------------------
@pytest.fixture(scope="session")
def admin_token() -> str:
    r = requests.post(f"{API}/login", json={"username": "admin", "password": "admin123"})
    assert r.status_code == 200, r.text
    data = r.json()
    assert data.get("user", {}).get("role") == "admin"
    return data.get("access_token") or data.get("token")


@pytest.fixture(scope="session")
def user_token() -> str:
    r = requests.post(f"{API}/login", json={"username": "user", "password": "user123"})
    assert r.status_code == 200, r.text
    d = r.json()
    return d.get("access_token") or d.get("token")


@pytest.fixture(scope="session")
def admin_headers(admin_token) -> dict:
    return {"Authorization": f"Bearer {admin_token}"}


@pytest.fixture(scope="session")
def mongo():
    client = MongoClient(MONGO_URL)
    yield client[DB_NAME]
    client.close()


# ------------------------------ health -------------------------------------
class TestHealth:
    def test_root(self):
        r = requests.get(f"{API}/")
        assert r.status_code == 200
        assert "message" in r.json()

    def test_admin_login_hidden_route(self):
        r = requests.post(f"{API}/admin/login", json={"username": "admin", "password": "admin123"})
        assert r.status_code == 200
        assert r.json()["user"]["role"] == "admin"


# --------------------------- customer sync engine --------------------------
class TestSyncEngine:
    """Local DB is source of truth. Deletes must NOT reappear."""

    def _make_customer(self, admin_headers) -> dict:
        code = f"TEST-{uuid4().hex[:8].upper()}"
        payload = {
            "customer_code": code, "customer_name": "TEST Customer",
            "segment": "Retail", "purchasing_size": "Small", "area": "Jakarta",
            "status": "Active", "bad_debt": False, "bad_debt_nominal": 0,
            "address": "TEST", "latitude": None, "longitude": None,
            "phone": "0", "whatsapp": "0", "pic_name": "T",
            "payment_terms": "Cash", "credit_limit": 0,
        }
        r = requests.post(f"{API}/customers", headers=admin_headers, json=payload)
        assert r.status_code == 201, r.text
        return r.json()

    def test_create_customer_creates_queue_item(self, admin_headers, mongo):
        cust = self._make_customer(admin_headers)
        cid = cust["id"]
        # allow the awaited enqueue to fully complete (worker may drain fast if remote is healthy)
        time.sleep(0.5)
        item = mongo.sync_queue.find_one({"entity_type": "customer", "entity_id": cid})
        state = mongo.sync_state.find_one({"entity_type": "customer", "entity_id": cid})
        assert item is not None or (state and state.get("synced_ever")), \
            "sync_queue must contain the new customer (or it was already pushed to Supabase)"
        # cleanup
        requests.delete(f"{API}/customers/{cid}", headers=admin_headers)

    def test_delete_customer_does_not_reappear(self, admin_headers, mongo):
        cust = self._make_customer(admin_headers)
        cid = cust["id"]
        r = requests.delete(f"{API}/customers/{cid}", headers=admin_headers)
        assert r.status_code == 200, r.text
        listing = requests.get(f"{API}/customers", headers=admin_headers).json()
        assert not any(c["id"] == cid for c in listing), "deleted customer must not be visible"
        # And if it was already synced remotely the queue should carry a tombstone,
        # otherwise the queue item is cancelled entirely.
        time.sleep(0.5)
        item = mongo.sync_queue.find_one({"entity_type": "customer", "entity_id": cid})
        state = mongo.sync_state.find_one({"entity_type": "customer", "entity_id": cid})
        if state and state.get("synced_ever"):
            assert item is not None and item["operation"] == "delete"
            assert item["payload"].get("deleted_at") is not None
        else:
            assert item is None, "CREATE->DELETE before-sync must be cancelled (no queue item)"

    def test_create_then_delete_cancels_when_sync_off(self, admin_headers, mongo):
        # disable sync (Supabase must be configured for this endpoint to succeed)
        toggle = requests.post(
            f"{API}/admin/supabase/sync-toggle", headers=admin_headers, json={"enabled": False}
        )
        if toggle.status_code == 400:
            pytest.skip("Supabase not configured in this env; cancel semantics still verified by test_delete_customer_does_not_reappear")
        assert toggle.status_code == 200, toggle.text
        try:
            cust = self._make_customer(admin_headers)
            cid = cust["id"]
            r = requests.delete(f"{API}/customers/{cid}", headers=admin_headers)
            assert r.status_code == 200
            time.sleep(0.5)
            item = mongo.sync_queue.find_one({"entity_type": "customer", "entity_id": cid})
            state = mongo.sync_state.find_one({"entity_type": "customer", "entity_id": cid})
            # If nothing was ever synced remotely, the cancel must remove the queue entry.
            if not (state and state.get("synced_ever")):
                assert item is None, "CREATE->DELETE before sync must cancel the queue item"
        finally:
            requests.post(
                f"{API}/admin/supabase/sync-toggle", headers=admin_headers, json={"enabled": True}
            )

    def test_offline_queue_pending_counter(self, admin_headers, mongo):
        toggle = requests.post(
            f"{API}/admin/supabase/sync-toggle", headers=admin_headers, json={"enabled": False}
        )
        if toggle.status_code == 400:
            pytest.skip("Supabase not configured; skipping toggle-based offline test")
        assert toggle.status_code == 200
        cid = None
        try:
            cust = self._make_customer(admin_headers)
            cid = cust["id"]
            time.sleep(0.5)
            status = requests.get(f"{API}/admin/supabase", headers=admin_headers).json()
            assert status.get("queue_pending", 0) >= 1
        finally:
            if cid:
                requests.delete(f"{API}/customers/{cid}", headers=admin_headers)
            requests.post(
                f"{API}/admin/supabase/sync-toggle", headers=admin_headers, json={"enabled": True}
            )


# ------------------------------- master CRUD -------------------------------
@pytest.mark.parametrize("entity", ["purchasing_size", "segment", "top"])
class TestMasterCRUD:
    def test_master_flow(self, admin_headers, entity):
        base = f"{API}/admin/master/{entity}"
        # LIST
        r = requests.get(base, headers=admin_headers)
        assert r.status_code == 200
        assert isinstance(r.json(), list)

        # CREATE
        name = f"TEST-{entity}-{uuid4().hex[:6]}"
        r = requests.post(base, headers=admin_headers, json={"name": name, "description": "d"})
        assert r.status_code == 201, r.text
        mid = r.json()["id"]

        # duplicate CREATE -> 409
        dup = requests.post(base, headers=admin_headers, json={"name": name, "description": "x"})
        assert dup.status_code == 409

        # RENAME
        new_name = name + "-renamed"
        upd = requests.put(f"{base}/{mid}", headers=admin_headers, json={"name": new_name, "description": "d2"})
        assert upd.status_code == 200
        assert upd.json()["name"] == new_name

        # DELETE
        d = requests.delete(f"{base}/{mid}", headers=admin_headers)
        assert d.status_code == 200
        # missing after delete
        listing = requests.get(base, headers=admin_headers).json()
        assert all(item["id"] != mid for item in listing)


class TestMasterUnknown:
    def test_unknown_master_type_404(self, admin_headers):
        r = requests.get(f"{API}/admin/master/category", headers=admin_headers)
        assert r.status_code == 404


# ------------------------------- user CRUD ---------------------------------
class TestUserCRUD:
    def test_create_and_login(self, admin_headers):
        uname = f"test_{uuid4().hex[:8]}"
        pw = "Passw0rd!"
        r = requests.post(
            f"{API}/admin/users", headers=admin_headers,
            json={"username": uname, "password": pw, "role": "user", "name": "Test User"},
        )
        assert r.status_code == 201, r.text
        body = r.json()
        for banned in ("password", "password_hash", "hashed_password"):
            assert banned not in body, f"user response must not expose {banned}"

        # login still works
        li = requests.post(f"{API}/login", json={"username": uname, "password": pw})
        assert li.status_code == 200
        assert li.json()["user"]["username"] == uname

        # cleanup
        requests.delete(f"{API}/admin/users/{uname}", headers=admin_headers)


# ------------------------------- logo upload -------------------------------
_PNG_1x1 = (
    b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01"
    b"\x08\x02\x00\x00\x00\x90wS\xde\x00\x00\x00\x0cIDATx\x9cc\xf8\xff\xff?\x00"
    b"\x05\xfe\x02\xfe\xa15\xf7X\x00\x00\x00\x00IEND\xaeB`\x82"
)


class TestLogoUpload:
    def test_upload_and_serve(self, admin_headers):
        files = {"file": ("logo.png", _PNG_1x1, "image/png")}
        r = requests.post(f"{API}/admin/upload-logo", headers=admin_headers, files=files)
        assert r.status_code == 200, r.text
        logo_url = r.json().get("logo_url")
        assert logo_url and logo_url.startswith("/api/files/"), logo_url
        served = requests.get(f"{BASE}{logo_url}")
        assert served.status_code == 200
        assert served.headers.get("content-type", "").startswith("image/")


# ------------------------------- Excel I/O ---------------------------------
class TestExcelIO:
    def test_template(self, admin_headers):
        r = requests.get(f"{API}/admin/template.xlsx", headers=admin_headers)
        assert r.status_code == 200, r.text
        assert "spreadsheet" in r.headers.get("content-type", "")
        wb = load_workbook(io.BytesIO(r.content), read_only=True)
        assert set(wb.sheetnames) >= {"Customers", "Purchasing Size", "Segment", "TOP"}

    def test_export_has_seeded_customers(self, admin_headers):
        r = requests.get(f"{API}/admin/export.xlsx", headers=admin_headers)
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content), read_only=True)
        ws = wb["Customers"]
        rows = list(ws.iter_rows(values_only=True))
        # header + at least 48 seeded customers
        data_rows = [row for row in rows[1:] if any(row)]
        assert len(data_rows) >= 48, f"expected >=48 customer data rows, got {len(data_rows)}"
        # master sheets should be populated
        for title in ("Purchasing Size", "Segment", "TOP"):
            mws = wb[title]
            master_rows = [row for row in list(mws.iter_rows(values_only=True))[1:] if any(row)]
            assert master_rows, f"master sheet '{title}' should not be empty"

    def _build_import_workbook(self, export_bytes: bytes):
        """Round-trip file: modify 1 phone, add 1 NEW customer w/ new segment, add 1 invalid row."""
        wb = load_workbook(io.BytesIO(export_bytes))
        ws = wb["Customers"]
        headers = [c.value for c in ws[1]]
        phone_idx = headers.index("Phone") + 1
        # modify phone on row 2 (first data row) with a unique value to guarantee change
        first_row_ref = ws.cell(row=2, column=1).value
        ws.cell(row=2, column=phone_idx, value=f"CHG-{uuid4().hex[:6]}")
        # Build a NEW customer row referencing a brand-new segment
        new_seg = f"TEST-SEG-{uuid4().hex[:6]}"
        new_code = f"TEST-NEW-{uuid4().hex[:6]}"
        row_vals = {
            "Customer Code": new_code, "Customer Name": "TEST New", "Segment": new_seg,
            "Purchasing Size": "Small", "Area": "Jakarta", "Status": "Active",
            "Bad Debt": "No", "Bad Debt Nominal": 0, "Address": "-", "Latitude": "",
            "Longitude": "", "Phone": "1", "WhatsApp": "1", "PIC Name": "-",
            "Payment Terms": "Cash", "Credit Limit": 0,
        }
        ws.append([row_vals.get(h, "") for h in headers])
        # Invalid row: bad Status
        invalid = dict(row_vals)
        invalid["Customer Code"] = f"TEST-BAD-{uuid4().hex[:6]}"
        invalid["Status"] = "Bogus"
        ws.append([invalid.get(h, "") for h in headers])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue(), new_code, invalid["Customer Code"], new_seg, first_row_ref

    def test_import_round_trip(self, admin_headers, mongo):
        exp = requests.get(f"{API}/admin/export.xlsx", headers=admin_headers)
        assert exp.status_code == 200
        payload, new_code, bad_code, new_seg, first_row_ref = self._build_import_workbook(exp.content)

        files = {"file": ("import.xlsx", payload,
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}

        # PREVIEW
        prev = requests.post(f"{API}/admin/import/xlsx/preview", headers=admin_headers, files=files)
        assert prev.status_code == 200, prev.text
        counts = prev.json()["customers"]["counts"]
        assert counts["create"] == 1, counts
        assert counts["update"] == 1, counts
        assert counts["error"] == 1, counts
        assert counts["skip"] >= 46, counts

        # COMMIT
        files = {"file": ("import.xlsx", payload,
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        com = requests.post(f"{API}/admin/import/xlsx/commit", headers=admin_headers, files=files)
        assert com.status_code == 200, com.text
        body = com.json()
        c = body["customers"]
        assert c["created"] == 1 and c["updated"] == 1 and c["failed"] == 1
        assert c["skipped"] >= 46

        # Verify persistence: new customer visible, no duplicates, master auto-created
        listing = requests.get(f"{API}/customers", headers=admin_headers).json()
        matches = [c for c in listing if c["customer_code"].lower() == new_code.lower()]
        assert len(matches) == 1, "no duplicates expected"

        segs = requests.get(f"{API}/admin/master/segment", headers=admin_headers).json()
        assert any(s["name"] == new_seg for s in segs), "referenced segment must be auto-created"

        # Cleanup: delete our TEST- new customer + auto-created segment
        cid = matches[0]["id"]
        requests.delete(f"{API}/customers/{cid}", headers=admin_headers)
        for s in segs:
            if s["name"] == new_seg:
                requests.delete(f"{API}/admin/master/segment/{s['id']}", headers=admin_headers)


if __name__ == "__main__":  # pragma: no cover
    import sys
    sys.exit(pytest.main([__file__, "-v", "--tb=short"]))
