"""Regression tests for the new features on iteration 4:
 * Trash (customer + area master): soft-delete -> trash -> restore / purge / empty
 * Trash counts include 5 entities (customer + 4 masters incl. area)
 * Excel export/template xlsx endpoints return 200
 * Import xlsx auto-creates referenced master values (incl. area)
 * GET /api/sync/status returns the expected shape/enum for user + admin
 * POST /api/admin/supabase/pull returns per-entity summary with expected keys
"""
from __future__ import annotations

import io
import os
from uuid import uuid4

import pytest
import requests
from openpyxl import load_workbook

BASE = os.environ["EXPO_PUBLIC_BACKEND_URL"].rstrip("/")
API = f"{BASE}/api"

VALID_STATUSES = {"offline", "syncing", "synced", "sync_failed", "conflict"}


# ----------------------------- fixtures ------------------------------------
@pytest.fixture(scope="session")
def admin_token() -> str:
    r = requests.post(f"{API}/admin/login", json={"username": "admin", "password": "admin123"})
    assert r.status_code == 200, r.text
    return r.json()["access_token"]


@pytest.fixture(scope="session")
def user_token() -> str:
    r = requests.post(f"{API}/login", json={"username": "user", "password": "user123"})
    assert r.status_code == 200, r.text
    return r.json()["access_token"]


@pytest.fixture(scope="session")
def admin_h(admin_token) -> dict:
    return {"Authorization": f"Bearer {admin_token}"}


@pytest.fixture(scope="session")
def user_h(user_token) -> dict:
    return {"Authorization": f"Bearer {user_token}"}


# ----------------------------- sync/status ---------------------------------
class TestSyncStatus:
    def test_shape_and_enum_admin(self, admin_h):
        r = requests.get(f"{API}/sync/status", headers=admin_h)
        assert r.status_code == 200, r.text
        b = r.json()
        for k in ("online", "status", "pending", "failed", "conflicts",
                  "last_sync_at", "last_pull_at"):
            assert k in b, f"missing key {k}"
        assert b["status"] in VALID_STATUSES
        assert isinstance(b["online"], bool)
        assert isinstance(b["pending"], int)
        assert isinstance(b["failed"], int)
        assert isinstance(b["conflicts"], int)

    def test_available_to_user(self, user_h):
        r = requests.get(f"{API}/sync/status", headers=user_h)
        assert r.status_code == 200
        assert r.json()["status"] in VALID_STATUSES

    def test_unauth(self):
        r = requests.get(f"{API}/sync/status")
        assert r.status_code == 401


# ----------------------------- trash counts --------------------------------
class TestTrashCounts:
    def test_counts_contain_five_entities(self, admin_h):
        r = requests.get(f"{API}/admin/trash-counts", headers=admin_h)
        assert r.status_code == 200
        b = r.json()
        expected = {"customer", "purchasing_size", "segment", "top", "area", "total"}
        assert expected <= set(b.keys()), f"missing keys {expected - set(b.keys())}"
        for k in ("customer", "purchasing_size", "segment", "top", "area"):
            assert isinstance(b[k], int)


# ----------------------------- trash customer flow -------------------------
class TestTrashCustomer:
    def _make_customer(self, admin_h) -> dict:
        code = f"TEST-{uuid4().hex[:8].upper()}"
        payload = {
            "customer_code": code, "customer_name": "TEST TrashCust",
            "segment": "Retail", "purchasing_size": "Small", "area": "Jakarta",
            "status": "Active", "bad_debt": False, "bad_debt_nominal": 0,
            "address": "TEST", "latitude": None, "longitude": None,
            "phone": "0", "whatsapp": "0", "pic_name": "T",
            "payment_terms": "Cash", "credit_limit": 0,
        }
        r = requests.post(f"{API}/customers", headers=admin_h, json=payload)
        assert r.status_code == 201, r.text
        return r.json()

    def test_soft_delete_restore(self, admin_h):
        cust = self._make_customer(admin_h)
        cid = cust["id"]
        # delete
        r = requests.delete(f"{API}/customers/{cid}", headers=admin_h)
        assert r.status_code == 200
        # appears in trash
        trash = requests.get(f"{API}/admin/trash/customer", headers=admin_h).json()
        assert any(t["id"] == cid for t in trash), "deleted customer must appear in trash"
        # restore
        r = requests.post(f"{API}/admin/trash/customer/restore",
                          headers=admin_h, json={"ids": [cid]})
        assert r.status_code == 200 and r.json()["restored"] == 1
        # back in main list
        listing = requests.get(f"{API}/customers", headers=admin_h).json()
        assert any(c["id"] == cid for c in listing), "restored customer must be visible"
        # cleanup: delete + purge
        requests.delete(f"{API}/customers/{cid}", headers=admin_h)
        requests.post(f"{API}/admin/trash/customer/purge",
                      headers=admin_h, json={"ids": [cid]})

    def test_purge_removes_permanently(self, admin_h):
        cust = self._make_customer(admin_h)
        cid = cust["id"]
        requests.delete(f"{API}/customers/{cid}", headers=admin_h)
        r = requests.post(f"{API}/admin/trash/customer/purge",
                          headers=admin_h, json={"ids": [cid]})
        assert r.status_code == 200 and r.json()["purged"] == 1
        # not in trash anymore
        trash = requests.get(f"{API}/admin/trash/customer", headers=admin_h).json()
        assert not any(t["id"] == cid for t in trash), "purged customer must NOT be in trash"
        # not in main list
        listing = requests.get(f"{API}/customers", headers=admin_h).json()
        assert not any(c["id"] == cid for c in listing)


# ----------------------------- trash area master flow ---------------------
class TestTrashAreaMaster:
    def test_area_soft_delete_and_restore(self, admin_h):
        name = f"TEST_Area_{uuid4().hex[:8]}"
        r = requests.post(f"{API}/admin/master/area", headers=admin_h,
                          json={"name": name, "description": "d"})
        assert r.status_code == 201, r.text
        aid = r.json()["id"]
        # delete
        d = requests.delete(f"{API}/admin/master/area/{aid}", headers=admin_h)
        assert d.status_code == 200
        # appears in area trash tab
        trash = requests.get(f"{API}/admin/trash/area", headers=admin_h).json()
        assert any(t["id"] == aid for t in trash), "deleted area must appear in trash area tab"
        # restore
        r = requests.post(f"{API}/admin/trash/area/restore",
                          headers=admin_h, json={"ids": [aid]})
        assert r.status_code == 200 and r.json()["restored"] == 1
        # active again
        areas = requests.get(f"{API}/admin/master/area", headers=admin_h).json()
        assert any(a["id"] == aid for a in areas)
        # cleanup
        requests.delete(f"{API}/admin/master/area/{aid}", headers=admin_h)
        requests.post(f"{API}/admin/trash/area/purge",
                      headers=admin_h, json={"ids": [aid]})


# ----------------------------- excel exports ------------------------------
class TestExcelEndpoints:
    def test_export_xlsx_200(self, admin_h):
        r = requests.get(f"{API}/admin/export.xlsx", headers=admin_h)
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content), read_only=True)
        assert "Customers" in wb.sheetnames
        # Area sheet must exist for round-trip with master
        assert "Area" in wb.sheetnames, f"expected 'Area' sheet, got {wb.sheetnames}"

    def test_template_xlsx_200(self, admin_h):
        r = requests.get(f"{API}/admin/template.xlsx", headers=admin_h)
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content), read_only=True)
        assert {"Customers", "Purchasing Size", "Segment", "TOP", "Area"} <= set(wb.sheetnames)


# --------------------------- import auto-master --------------------------
class TestImportAutoArea:
    def test_import_auto_creates_area(self, admin_h):
        # start from the current export so schema+existing rows match
        exp = requests.get(f"{API}/admin/export.xlsx", headers=admin_h)
        assert exp.status_code == 200
        wb = load_workbook(io.BytesIO(exp.content))
        ws = wb["Customers"]
        headers = [c.value for c in ws[1]]
        new_area = f"TEST_AREA_{uuid4().hex[:6]}"
        new_code = f"TEST-AREA-NEW-{uuid4().hex[:6]}"
        row_vals = {
            "Customer Code": new_code, "Customer Name": "TEST Area Import",
            "Segment": "Retail", "Purchasing Size": "Small", "Area": new_area,
            "Status": "Active", "Bad Debt": "No", "Bad Debt Nominal": 0,
            "Address": "-", "Latitude": "", "Longitude": "", "Phone": "1",
            "WhatsApp": "1", "PIC Name": "-", "Payment Terms": "Cash", "Credit Limit": 0,
        }
        ws.append([row_vals.get(h, "") for h in headers])
        buf = io.BytesIO(); wb.save(buf)
        payload = buf.getvalue()

        files = {"file": ("import.xlsx", payload,
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        com = requests.post(f"{API}/admin/import/xlsx/commit", headers=admin_h, files=files)
        assert com.status_code == 200, com.text
        body = com.json()
        # master summary must include 'area' (not 'undefined')
        assert "master" in body and "area" in body["master"], f"missing area in {body.get('master')}"
        # the area referenced by the new customer should be auto-created
        areas = requests.get(f"{API}/admin/master/area", headers=admin_h).json()
        matching = [a for a in areas if a["name"] == new_area]
        assert matching, f"area '{new_area}' must be auto-created on import"

        # cleanup
        listing = requests.get(f"{API}/customers", headers=admin_h).json()
        for c in listing:
            if c["customer_code"].lower() == new_code.lower():
                requests.delete(f"{API}/customers/{c['id']}", headers=admin_h)
                requests.post(f"{API}/admin/trash/customer/purge",
                              headers=admin_h, json={"ids": [c["id"]]})
        for a in matching:
            requests.delete(f"{API}/admin/master/area/{a['id']}", headers=admin_h)
            requests.post(f"{API}/admin/trash/area/purge",
                          headers=admin_h, json={"ids": [a['id']]})


# --------------------------- supabase pull -------------------------------
class TestSupabasePull:
    def test_pull_summary_shape(self, admin_h):
        r = requests.post(f"{API}/admin/supabase/pull", headers=admin_h)
        # Either not connected (400) or returns a summary
        if r.status_code == 400:
            pytest.skip("Supabase not connected in this env")
        assert r.status_code == 200, r.text
        b = r.json()
        assert b.get("success") is True
        assert "summary" in b
        s = b["summary"]
        # customer must have full shape
        assert "customer" in s
        for k in ("created", "updated", "deleted", "skipped", "conflicts"):
            assert k in s["customer"], f"customer summary missing {k}"
        # masters: either same shape or an {error:...} object (e.g., area table missing)
        for e in ("purchasing_size", "segment", "top", "area"):
            assert e in s, f"summary missing entity {e}"
            entry = s[e]
            assert "error" in entry or "created" in entry, entry

    def test_pull_does_not_resurrect_purged(self, admin_h):
        # Create+delete+purge a customer, then pull; the row must not come back.
        code = f"TEST-PURGE-{uuid4().hex[:6]}"
        payload = {
            "customer_code": code, "customer_name": "TEST Purge", "segment": "Retail",
            "purchasing_size": "Small", "area": "Jakarta", "status": "Active",
            "bad_debt": False, "bad_debt_nominal": 0, "address": "-", "latitude": None,
            "longitude": None, "phone": "0", "whatsapp": "0", "pic_name": "T",
            "payment_terms": "Cash", "credit_limit": 0,
        }
        r = requests.post(f"{API}/customers", headers=admin_h, json=payload)
        assert r.status_code == 201
        cid = r.json()["id"]
        requests.delete(f"{API}/customers/{cid}", headers=admin_h)
        requests.post(f"{API}/admin/trash/customer/purge",
                      headers=admin_h, json={"ids": [cid]})
        pull = requests.post(f"{API}/admin/supabase/pull", headers=admin_h)
        if pull.status_code == 400:
            pytest.skip("Supabase not connected")
        listing = requests.get(f"{API}/customers", headers=admin_h).json()
        assert not any(c["id"] == cid for c in listing), "purged customer must not be resurrected"


if __name__ == "__main__":
    import sys
    sys.exit(pytest.main([__file__, "-v", "--tb=short"]))
